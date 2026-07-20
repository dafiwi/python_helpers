#!/usr/bin/env python3
# -*- coding: utf-8 -*-


import csv
from importlib.metadata import distributions
import locale
import mysql.connector
from openpyxl import load_workbook, Workbook
import os
from pathlib import Path
import platform
import sys
import time


# Internal Functions

def enter_filename():
    filename = input("\nPlease enter the whole filename: ")
    return filename


def enter_text():
    text = input("\nPlease enter a text: ")
    return text


def enter_column():
    column = input("Please enter the column name: ")
    return column


def load_xlsx_file():
    filename = enter_filename()
    workbook = load_workbook(filename)
    worksheet = workbook.active
    return filename, workbook, worksheet


def load_xlsx_file_and_choose_column():
    filename = enter_filename()
    workbook = load_workbook(filename)
    worksheet = workbook.active
    column = enter_column()
    return filename, workbook, worksheet, column


def convert_bytes_to_mebibytes(size_in_bytes):
    size_in_mebibytes = size_in_bytes / 1024 / 1024
    return size_in_mebibytes



# Helper Functions

def remove_eur_appendix_in_a_xlsx_file():
    filename, workbook, worksheet, column = load_xlsx_file_and_choose_column()

    for cell in worksheet[column]:
        if cell.value is not None:
            cell.value = str(cell.value).replace("EUR", "").strip()

    workbook.save(filename)


def remove_usd_appendix_in_a_xlsx_file():
    filename, workbook, worksheet, column = load_xlsx_file_and_choose_column()

    for cell in worksheet[column]:
        if cell.value is not None:
            cell.value = str(cell.value).replace("USD", "").strip()

    workbook.save(filename)


def replace_comma_with_dot_in_a_xlsx_file():
    filename, workbook, worksheet, column = load_xlsx_file_and_choose_column()

    for cell in worksheet[column]:
        if cell.value is not None:
            cell.value = str(cell.value).replace(",", ".")

    workbook.save(filename)


def remove_percentage_sign_in_a_xlsx_file():
    filename, workbook, worksheet, column = load_xlsx_file_and_choose_column()

    for cell in worksheet[column]:
        if cell.value is not None:
            cell.value = str(cell.value).replace("%", "")

    workbook.save(filename)


def convert_thousands_in_whole_numbers_in_a_xlsx_file():
    filename, workbook, worksheet, column = load_xlsx_file_and_choose_column()

    for cell in worksheet[column]:
        if cell.value is not None and any(c.isdigit() for c in str(cell.value)) and "K" in str(cell.value):
            cell.value = str(cell.value).replace("K", "").strip()
            cell.value = float(cell.value)
            cell.value *= 1_000

    workbook.save(filename)


def convert_millions_in_whole_numbers_in_a_xlsx_file():
    filename, workbook, worksheet, column = load_xlsx_file_and_choose_column()

    for cell in worksheet[column]:
        if cell.value is not None and any(c.isdigit() for c in str(cell.value)) and "M" in str(cell.value):
            cell.value = str(cell.value).replace("M", "").strip()
            cell.value = float(cell.value)
            cell.value *= 1_000_000

    workbook.save(filename)


def convert_billions_in_whole_numbers_in_a_xlsx_file():
    filename, workbook, worksheet, column = load_xlsx_file_and_choose_column()

    for cell in worksheet[column]:
        if cell.value is not None and any(c.isdigit() for c in str(cell.value)) and "B" in str(cell.value):
            cell.value = str(cell.value).replace("B", "").strip()
            cell.value = float(cell.value)
            cell.value *= 1_000_000_000

    workbook.save(filename)


def remove_all_dots_but_the_rightmost_dot_in_a_number():
    number = input("\nPlease enter a number with multiple dots in between: ")
    if number.count(".") > 1:
        left_part, right_part = number.rsplit(".", 1)
        number = left_part.replace(".", "") + "." + right_part
    number = float(number)
    print(f"This is the number with only the rightmost decimal point remaining: {number}")
    return number


def convert_character_to_unicode_value():
    character = input("\nPlease enter a character: ")
    unicode_value = ord(character)
    print(f"The character {character} has the unicode value {unicode_value}")
    return unicode_value


def convert_unicode_value_to_character():
    unicode_value = input("\nPlease enter a unicode value: ")
    character = chr(int(unicode_value))
    print(f"The unicode value {unicode_value} represents the character {character}")
    return character


def convert_float_to_percentage_value():
    float_value = float(input("\nPlease enter a float value: "))
    percentage_value = float_value * 100
    print(f"The float value {float_value} corresponds to the percentage value {percentage_value} %.")
    return percentage_value


def write_text_into_a_txt_file():
    text = enter_text() + "\n"
    filename = input("Please enter a filename (without extension): ") + ".txt"
    with open(filename, "w", encoding="utf-8") as file:
        file.write(text)
    print(f"The text file {filename} was created successfully.")


def append_text_to_a_txt_file():
    filename = enter_filename()
    text = enter_text() + "\n"
    with open(filename, "a", encoding="utf-8") as file:
        file.write(text)
    print(f"Some text was appended to the file {filename} successfully.")


def read_a_txt_file():
    filename = enter_filename()
    with open(filename, "r", encoding="utf-8") as file:
        content = file.read()
    print("This is the content of the file: \n")
    print(content)


def check_if_a_string_contains_a_number():
    string_input = input("\nPlease enter a string: ")
    if any(character.isdigit() for character in str(string_input)):
        print(f"{string_input} contains a number.")
        return True
    else:
        print(f"{string_input} does not contain a number.")
        return False


def check_if_a_floating_point_number_is_a_whole_number():
    floating_point_number_input = float(input("\nPlease enter a floating point number: "))
    if floating_point_number_input.is_integer():
        print(f"{floating_point_number_input} is a whole number.")
        return True
    else:
        print(f"{floating_point_number_input} is not a whole number.")
        return False


def system_info():
    platform.invalidate_caches()

    print("\nSystem Information\n")
    print(f"Computer: {platform.machine()}")
    print(f"Processor: {platform.processor()}")
    print(f"Architecture: {platform.architecture()[0]}")
    print(f"System: {platform.system()}")
    print(f"Release: {platform.release()}")
    print(f"Version: {platform.version()}")
    print(f"Network: {platform.node()}")


def python_info():
    print("\nPython Information\n")
    print(f"Python Version: {platform.python_version()}")
    print(f"Python Revision: {platform.python_revision()}")
    print(f"Python Build Date: {platform.python_build()[1]}")
    print(f"Python Compiler: {platform.python_compiler()}")
    print(f"Python Implementation: {platform.python_implementation()}")


def get_directory_size():
    directory_path = input("\nPlease enter the directory path: ")
    directory_size = 0
    for root, _, files in os.walk(directory_path):
        for file in files:
            file_path = os.path.join(root, file)
            try:
                directory_size += os.path.getsize(file_path)
            except OSError as exception:
                print(exception)
    print(f"{directory_path} needs {convert_bytes_to_mebibytes(directory_size):7.2f} MB disk space.")

    return directory_size


def print_storage_size_for_whole_installed_python_modules():
    print("\nInstalled Python Modules:\n")
    for distribution in sorted(distributions(), key=lambda d: d.metadata["Name"].lower()):
        storage_size = sum(
            path.stat().st_size
            for file in distribution.files or []
            if (path := Path(distribution.locate_file(file))).is_file()
        )
        print(f"{distribution.metadata['Name']:<25} {convert_bytes_to_mebibytes(storage_size):7.2f} MB")


def convert_xlsx_to_csv():
    print("\nPlease enter the name of the XLSX source file first.")
    filename, workbook, worksheet = load_xlsx_file()
    print("\nPlease enter the name of the CSV target file next.")
    target_file = enter_filename()

    with open(target_file, "w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)

        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(row)

    print(f"\n\033[32m{filename} has been successfully converted to the CSV file {target_file}.\033[0m\n")


def convert_csv_to_xlsx():
    print("\nPlease enter the name of the CSV source file first.")
    source_file = enter_filename()
    workbook = Workbook()
    worksheet = workbook.active

    with open(source_file, "r", encoding="utf-8", newline="") as csv_file:
        reader = csv.reader(csv_file, delimiter=",")

        for line in reader:
            worksheet.append(line)

    target_file = Path(source_file).with_suffix(".xlsx")
    if target_file.exists():
        print("\nPlease enter the name of the XLSX target file next.")
        target_file = enter_filename()
    workbook.save(target_file)

    print(f"\n\033[32m{source_file} has been successfully converted to the XLSX file {target_file}.\033[0m\n")


def read_data_from_a_csv_file():
    filename = enter_filename()
    with open(filename, newline="", encoding="utf-8") as file:
        reader = csv.reader(file)
        for row in reader:
            print(row)


def get_file_path():
    file_path = Path(__file__).resolve()
    print(file_path)
    return file_path


def get_directory_path():
    directory_path = Path(__file__).resolve().parent
    print(directory_path)
    return directory_path


def connect_to_the_local_mysql_database():
    try:
        mysql_database_password = input("\nPlease enter your MySQL database password: ")
        connection = mysql.connector.connect(host="localhost", port=3306, user="root", password=mysql_database_password, use_pure=True)
        if connection.is_connected():
            print(f"\nYou are connected to your MySQL database. Version: {connection.server_info}")
            connection.close()
            print("The connection has been closed again. You should not leave the database connection open.")
    except mysql.connector.Error as error:
        print(f"\n{error}")
        print("Error: You are not connected to your MySQL database.")


def print_date_and_time_in_english():
    print(time.strftime("\nDate: %d.%m.%Y"))
    print(time.strftime("Time: %H:%M:%S"))
    print(time.strftime("Day of the week: %A"))


def print_date_and_time_in_german():
    locale.setlocale(locale.LC_TIME, 'de_DE')
    print(time.strftime("\nDatum: %d.%m.%Y"))
    print(time.strftime("Zeit: %H:%M:%S"))
    print(time.strftime("Wochentag: %A"))



functions = {
    "1": remove_eur_appendix_in_a_xlsx_file,
    "2": remove_usd_appendix_in_a_xlsx_file,
    "3": replace_comma_with_dot_in_a_xlsx_file,
    "4": remove_percentage_sign_in_a_xlsx_file,
    "5": convert_thousands_in_whole_numbers_in_a_xlsx_file,
    "6": convert_millions_in_whole_numbers_in_a_xlsx_file,
    "7": convert_billions_in_whole_numbers_in_a_xlsx_file,
    "8": remove_all_dots_but_the_rightmost_dot_in_a_number,
    "10": convert_character_to_unicode_value,
    "11": convert_unicode_value_to_character,
    "12": convert_float_to_percentage_value,
    "20": write_text_into_a_txt_file,
    "21": append_text_to_a_txt_file,
    "22": read_a_txt_file,
    "30": check_if_a_string_contains_a_number,
    "31": check_if_a_floating_point_number_is_a_whole_number,
    "40": system_info,
    "41": python_info,
    "42": get_directory_size,
    "43": print_storage_size_for_whole_installed_python_modules,
    "50": convert_xlsx_to_csv,
    "51": convert_csv_to_xlsx,
    "52": read_data_from_a_csv_file,
    "60": get_file_path,
    "61": get_directory_path,
    "70": connect_to_the_local_mysql_database,
    "90": print_date_and_time_in_english,
    "91": print_date_and_time_in_german,
}



def main():
    print("\nPython Helper Functions")

    print("\nAvailable functions:")
    for key, function in functions.items():
        print(f"{key}: {function.__name__}")

    function_choice = input("\nEnter the number of the function you want to execute (or QUIT to exit the program): ").strip()

    if function_choice == "QUIT":
        sys.exit()
    elif function_choice in functions:
        functions[function_choice]()
    else:
        print(function_choice + " does not exist.")

if __name__ == "__main__":
    main()
